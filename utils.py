"""
Módulo de utilidades
Funções auxiliares para formatação, validação e exportação
"""

from datetime import datetime, timedelta
from typing import List, Dict
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class Formatador:
    """Classe para formatação de dados"""
    
    @staticmethod
    def formatar_moeda(valor: float) -> str:
        """Formata valor para moeda brasileira"""
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
    @staticmethod
    def formatar_data(data_str: str, formato_entrada: str = "%Y-%m-%d %H:%M:%S",
                     formato_saida: str = "%d/%m/%Y") -> str:
        """Formata data para o padrão brasileiro"""
        try:
            if not data_str:
                return ""
            data = datetime.strptime(data_str, formato_entrada)
            return data.strftime(formato_saida)
        except:
            return data_str
    
    @staticmethod
    def formatar_data_hora(data_str: str) -> str:
        """Formata data e hora"""
        return Formatador.formatar_data(data_str, formato_saida="%d/%m/%Y %H:%M")
    
    @staticmethod
    def formatar_porcentagem(valor: float, decimais: int = 1) -> str:
        """Formata valor como porcentagem"""
        return f"{valor:.{decimais}f}%"
    
    @staticmethod
    def calcular_margem(preco_venda: float, preco_custo: float) -> float:
        """Calcula margem de lucro percentual"""
        if preco_custo == 0:
            return 0
        return ((preco_venda - preco_custo) / preco_custo) * 100
    
    @staticmethod
    def calcular_preco_com_margem(preco_custo: float, margem_percentual: float) -> float:
        """Calcula preço de venda baseado em margem desejada"""
        return preco_custo * (1 + (margem_percentual / 100))

class Validador:
    """Classe para validação de dados"""
    
    @staticmethod
    def validar_numero(valor: str) -> bool:
        """Valida se é um número válido"""
        try:
            float(valor.replace(",", "."))
            return True
        except:
            return False
    
    @staticmethod
    def validar_inteiro(valor: str) -> bool:
        """Valida se é um número inteiro válido"""
        try:
            int(valor)
            return True
        except:
            return False
    
    @staticmethod
    def validar_data(data_str: str, formato: str = "%Y-%m-%d") -> bool:
        """Valida se é uma data válida"""
        try:
            datetime.strptime(data_str, formato)
            return True
        except:
            return False
    
    @staticmethod
    def limpar_numero(valor: str) -> str:
        """Remove caracteres não numéricos mantendo ponto e vírgula"""
        return ''.join(c for c in valor if c.isdigit() or c in '.,')
    
    @staticmethod
    def converter_moeda_float(valor_str: str) -> float:
        """Converte string de moeda para float"""
        valor_limpo = valor_str.replace("R$", "").replace(" ", "")
        valor_limpo = valor_limpo.replace(".", "").replace(",", ".")
        try:
            return float(valor_limpo)
        except:
            return 0.0

class Periodo:
    """Classe para trabalhar com períodos de tempo"""
    
    @staticmethod
    def hoje() -> str:
        """Retorna a data de hoje"""
        return datetime.now().strftime("%Y-%m-%d")
    
    @staticmethod
    def ontem() -> str:
        """Retorna a data de ontem"""
        return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    
    @staticmethod
    def inicio_semana() -> str:
        """Retorna o início da semana atual (segunda-feira)"""
        hoje = datetime.now()
        inicio = hoje - timedelta(days=hoje.weekday())
        return inicio.strftime("%Y-%m-%d")
    
    @staticmethod
    def fim_semana() -> str:
        """Retorna o fim da semana atual (domingo)"""
        hoje = datetime.now()
        fim = hoje + timedelta(days=(6 - hoje.weekday()))
        return fim.strftime("%Y-%m-%d")
    
    @staticmethod
    def inicio_mes() -> str:
        """Retorna o primeiro dia do mês atual"""
        hoje = datetime.now()
        return hoje.replace(day=1).strftime("%Y-%m-%d")
    
    @staticmethod
    def fim_mes() -> str:
        """Retorna o último dia do mês atual"""
        hoje = datetime.now()
        proximo_mes = hoje.replace(day=28) + timedelta(days=4)
        ultimo_dia = proximo_mes - timedelta(days=proximo_mes.day)
        return ultimo_dia.strftime("%Y-%m-%d")
    
    @staticmethod
    def inicio_ano() -> str:
        """Retorna o primeiro dia do ano"""
        hoje = datetime.now()
        return hoje.replace(month=1, day=1).strftime("%Y-%m-%d")
    
    @staticmethod
    def fim_ano() -> str:
        """Retorna o último dia do ano"""
        hoje = datetime.now()
        return hoje.replace(month=12, day=31).strftime("%Y-%m-%d")
    
    @staticmethod
    def mes_anterior() -> tuple:
        """Retorna tupla (inicio, fim) do mês anterior"""
        hoje = datetime.now()
        primeiro_dia_mes_atual = hoje.replace(day=1)
        ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
        primeiro_dia_mes_anterior = ultimo_dia_mes_anterior.replace(day=1)
        
        return (
            primeiro_dia_mes_anterior.strftime("%Y-%m-%d"),
            ultimo_dia_mes_anterior.strftime("%Y-%m-%d")
        )
    
    @staticmethod
    def ultimos_n_dias(n: int) -> tuple:
        """Retorna tupla (inicio, fim) dos últimos N dias"""
        hoje = datetime.now()
        inicio = hoje - timedelta(days=n-1)
        return (inicio.strftime("%Y-%m-%d"), hoje.strftime("%Y-%m-%d"))

class ExportadorPDF:
    """Classe para exportação de relatórios em PDF"""
    
    @staticmethod
    def gerar_relatorio_vendas(vendas: List[Dict], arquivo: str, periodo: str = ""):
        """Gera relatório de vendas em PDF"""
        doc = SimpleDocTemplate(arquivo, pagesize=A4)
        elementos = []
        styles = getSampleStyleSheet()
        
        # Título
        titulo_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        titulo = Paragraph(f"Relatório de Vendas{' - ' + periodo if periodo else ''}", titulo_style)
        elementos.append(titulo)
        elementos.append(Spacer(1, 20))
        
        # Tabela de vendas
        if vendas:
            dados = [['ID', 'Produto', 'Qtd', 'Valor Unit.', 'Total', 'Data']]
            total_geral = 0
            
            for venda in vendas:
                dados.append([
                    str(venda['id']),
                    venda['produto_nome'][:30],
                    str(venda['quantidade']),
                    Formatador.formatar_moeda(venda['preco_unitario']),
                    Formatador.formatar_moeda(venda['valor_total']),
                    Formatador.formatar_data(venda['data_venda'])
                ])
                total_geral += venda['valor_total']
            
            # Linha de total
            dados.append(['', '', '', 'TOTAL:', Formatador.formatar_moeda(total_geral), ''])
            
            tabela = Table(dados, colWidths=[0.6*inch, 2.5*inch, 0.7*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            tabela.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8e8e8')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elementos.append(tabela)
        else:
            texto = Paragraph("Nenhuma venda registrada no período.", styles['Normal'])
            elementos.append(texto)
        
        doc.build(elementos)
    
    @staticmethod
    def gerar_relatorio_produtos(produtos: List[Dict], arquivo: str):
        """Gera relatório de produtos em PDF"""
        doc = SimpleDocTemplate(arquivo, pagesize=A4)
        elementos = []
        styles = getSampleStyleSheet()
        
        # Título
        titulo = Paragraph("Relatório de Produtos", styles['Title'])
        elementos.append(titulo)
        elementos.append(Spacer(1, 20))
        
        if produtos:
            dados = [['ID', 'Nome', 'Categoria', 'Estoque', 'P. Custo', 'P. Venda', 'Margem']]
            
            for produto in produtos:
                margem = Formatador.calcular_margem(produto['preco_venda'], produto['preco_custo'])
                dados.append([
                    str(produto['id']),
                    produto['nome'][:25],
                    produto.get('categoria_nome', 'S/Cat')[:15],
                    str(produto['estoque']),
                    Formatador.formatar_moeda(produto['preco_custo']),
                    Formatador.formatar_moeda(produto['preco_venda']),
                    Formatador.formatar_porcentagem(margem)
                ])
            
            tabela = Table(dados, colWidths=[0.5*inch, 2*inch, 1.3*inch, 0.8*inch, 1*inch, 1*inch, 0.8*inch])
            tabela.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ca02c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elementos.append(tabela)
        else:
            texto = Paragraph("Nenhum produto cadastrado.", styles['Normal'])
            elementos.append(texto)
        
        doc.build(elementos)
    
    @staticmethod
    def gerar_relatorio_financeiro(dados_financeiros: Dict, arquivo: str, periodo: str = ""):
        """Gera relatório financeiro em PDF"""
        doc = SimpleDocTemplate(arquivo, pagesize=A4)
        elementos = []
        styles = getSampleStyleSheet()
        
        # Título
        titulo = Paragraph(f"Relatório Financeiro{' - ' + periodo if periodo else ''}", styles['Title'])
        elementos.append(titulo)
        elementos.append(Spacer(1, 30))
        
        # Resumo financeiro
        dados_tabela = [
            ['Descrição', 'Valor'],
            ['Receita Total', Formatador.formatar_moeda(dados_financeiros.get('receita_total', 0))],
            ['Lucro Bruto', Formatador.formatar_moeda(dados_financeiros.get('lucro_bruto', 0))],
            ['Despesas', Formatador.formatar_moeda(dados_financeiros.get('despesas', 0))],
            ['Lucro Líquido', Formatador.formatar_moeda(dados_financeiros.get('lucro_liquido', 0))]
        ]
        
        tabela = Table(dados_tabela, colWidths=[3*inch, 2*inch])
        tabela.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ff7f0e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d4edda')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elementos.append(tabela)
        doc.build(elementos)

class ExportadorExcel:
    """Classe para exportação de relatórios em Excel"""
    
    @staticmethod
    def gerar_relatorio_vendas(vendas: List[Dict], arquivo: str):
        """Gera relatório de vendas em Excel"""
        df = pd.DataFrame(vendas)
        
        if not df.empty:
            # Formatar colunas
            colunas_renomear = {
                'id': 'ID',
                'produto_nome': 'Produto',
                'quantidade': 'Quantidade',
                'preco_unitario': 'Preço Unitário',
                'valor_total': 'Valor Total',
                'cliente': 'Cliente',
                'data_venda': 'Data da Venda'
            }
            df = df.rename(columns=colunas_renomear)
            
            # Selecionar apenas colunas relevantes
            colunas_manter = ['ID', 'Produto', 'Quantidade', 'Preço Unitário', 
                            'Valor Total', 'Cliente', 'Data da Venda']
            df = df[[col for col in colunas_manter if col in df.columns]]
        
        # Salvar em Excel
        with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Vendas', index=False)
            
            # Formatar planilha
            workbook = writer.book
            worksheet = writer.sheets['Vendas']
            
            # Estilo do cabeçalho
            header_fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def gerar_relatorio_produtos(produtos: List[Dict], arquivo: str):
        """Gera relatório de produtos em Excel"""
        df = pd.DataFrame(produtos)
        
        if not df.empty:
            # Adicionar coluna de margem
            df['margem'] = df.apply(
                lambda row: Formatador.calcular_margem(row['preco_venda'], row['preco_custo']),
                axis=1
            )
            
            colunas_renomear = {
                'id': 'ID',
                'nome': 'Nome',
                'categoria_nome': 'Categoria',
                'preco_custo': 'Preço Custo',
                'preco_venda': 'Preço Venda',
                'estoque': 'Estoque',
                'estoque_minimo': 'Estoque Mínimo',
                'margem': 'Margem (%)'
            }
            df = df.rename(columns=colunas_renomear)
            
            colunas_manter = ['ID', 'Nome', 'Categoria', 'Preço Custo', 'Preço Venda',
                            'Estoque', 'Estoque Mínimo', 'Margem (%)']
            df = df[[col for col in colunas_manter if col in df.columns]]
        
        with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Produtos', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Produtos']
            
            header_fill = PatternFill(start_color='2ca02c', end_color='2ca02c', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def gerar_relatorio_completo(dados: Dict, arquivo: str):
        """Gera relatório completo com múltiplas abas"""
        with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
            # Aba de vendas
            if 'vendas' in dados and dados['vendas']:
                df_vendas = pd.DataFrame(dados['vendas'])
                df_vendas.to_excel(writer, sheet_name='Vendas', index=False)
            
            # Aba de produtos
            if 'produtos' in dados and dados['produtos']:
                df_produtos = pd.DataFrame(dados['produtos'])
                df_produtos.to_excel(writer, sheet_name='Produtos', index=False)
            
            # Aba de despesas
            if 'despesas' in dados and dados['despesas']:
                df_despesas = pd.DataFrame(dados['despesas'])
                df_despesas.to_excel(writer, sheet_name='Despesas', index=False)
            
            # Aba de resumo
            if 'resumo' in dados:
                df_resumo = pd.DataFrame([dados['resumo']])
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
